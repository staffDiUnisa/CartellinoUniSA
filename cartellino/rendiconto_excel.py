from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from cartellino.timesheet_progetto import ConfigTimesheet

_MONTH_IT_TITLE = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
]

_MONTH_IT_UPPER = [m.upper() for m in _MONTH_IT_TITLE]

_MONTH_IT_TO_NUM: dict[str, int] = {m: i + 1 for i, m in enumerate(_MONTH_IT_TITLE)}

# Fill weekend (sabato/domenica): grigio chiaro tema, identico al template
_WEEKEND_FILL = PatternFill(
    fill_type="solid",
    fgColor=Color(theme=0, tint=-0.1499984740745262),
)
_NO_FILL = PatternFill(fill_type=None)

# Righe della griglia attività
_ROW_ORE_PROGETTO = 24
_ROW_ORE_ALTRI_PROG = 25
_ROW_ORE_ORDINARIE = 26
_ROW_ORE_ALTRO = 27

# Colonna del giorno 1 (C = indice 3, 1-based)
_COL_GIORNO_BASE = 3


def _col_for_day(day: int) -> int:
    """Indice di colonna (1-based) per il giorno del mese. Giorno 1 → C (3)."""
    return _COL_GIORNO_BASE + day - 1


def _normalizza_nome_foglio(nome: str) -> str:
    """Normalizza spazi multipli (es. 'Gennaio  2025' → 'Gennaio 2025')."""
    return re.sub(r"\s+", " ", nome).strip()


class RendicontoExcel:
    """Popola il template Excel di rendicontazione con i dati del timesheet."""

    def __init__(
        self,
        config: ConfigTimesheet,
        monthly_data: dict[int, pd.DataFrame],
        output_folder: Path,
    ) -> None:
        self._config = config
        self._monthly_data = monthly_data
        self._output_folder = output_folder

    def genera(self) -> Path:
        cfg = self._config
        assert cfg.template_rendiconto is not None, "template_rendiconto non impostato"

        wb = openpyxl.load_workbook(cfg.template_rendiconto)

        # Costruisce mappa {old_sheet_name: new_sheet_name} per i fogli mensili
        rename_map: dict[str, str] = {}
        for sheet_name in wb.sheetnames:
            norm = _normalizza_nome_foglio(sheet_name)
            parts = norm.split(" ", 1)
            if len(parts) < 2 or parts[0] not in _MONTH_IT_TO_NUM:
                continue
            month_name_it = parts[0]
            new_name = f"{month_name_it} {cfg.anno}"
            rename_map[sheet_name] = new_name

        # Processa ogni foglio mensile (rinomina + contenuto)
        for old_name, new_name in rename_map.items():
            ws = wb[old_name]
            month_num = _MONTH_IT_TO_NUM[new_name.split(" ")[0]]

            ws.title = new_name

            # Intestazione anno/mese
            ws["N12"] = _MONTH_IT_UPPER[month_num - 1]
            ws["AG12"] = cfg.anno
            ws["C22"] = f"Mese di {_MONTH_IT_TITLE[month_num - 1]} {cfg.anno}"

            # Metadati progetto
            ws["C15"] = cfg.cup
            ws["C16"] = cfg.codice

            if cfg.persona:
                ws["C18"] = cfg.persona.figura_professionale
                ws["C19"] = cfg.persona.nome
                ws["Y19"] = cfg.persona.cognome
                ws["C20"] = cfg.persona.codice_fiscale

            # Ore totali del mese e griglia giornaliera
            mdf = self._monthly_data.get(month_num)
            ws["Y20"] = round(mdf["ore_progetto"].sum(), 4) if mdf is not None else None

            self._aggiorna_griglia(ws, month_num, mdf)

        # Aggiorna riferimenti nel foglio Riassuntivo (solo stringhe formula)
        if "Riassuntivo" in wb.sheetnames:
            ws_riass = wb["Riassuntivo"]
            for row in ws_riass.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = _aggiorna_riferimenti(cell.value, rename_map)

        # Salva
        output_folder = self._output_folder / "ore_svolte_per_giorno" / cfg.nome
        output_folder.mkdir(parents=True, exist_ok=True)

        if cfg.persona:
            filename = (
                f"TS_{cfg.nome}_{cfg.anno}"
                f"_{cfg.persona.cognome}_{cfg.persona.nome}.xlsx"
            )
        else:
            filename = f"TS_{cfg.nome}_{cfg.anno}.xlsx"

        output_path = output_folder / filename
        wb.save(output_path)
        print(f"Rendiconto generato: {output_path}")
        return output_path

    # ------------------------------------------------------------------

    def _aggiorna_griglia(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        month_num: int,
        mdf: pd.DataFrame | None,
    ) -> None:
        """Aggiorna colori weekend e ore nelle righe 24–27 per il mese dato."""
        cfg = self._config
        _, last_day = calendar.monthrange(cfg.anno, month_num)

        # Lookup day → (ore_progetto, ore_ordinarie)
        day_to_ore: dict[int, tuple[float, float]] = {}
        if mdf is not None:
            for _, row in mdf.iterrows():
                d = row["date_only"]
                day_num = d.day if isinstance(d, date) else d
                day_to_ore[day_num] = (float(row["ore_progetto"]), float(row["ore_ordinarie"]))

        for day in range(1, 32):
            col_idx = _col_for_day(day)
            col_letter = get_column_letter(col_idx)

            if day > last_day:
                # Giorno inesistente nel mese: rimuovi colore e valore
                for r in range(_ROW_ORE_PROGETTO, _ROW_ORE_ALTRO + 1):
                    cell = ws[f"{col_letter}{r}"]
                    cell.fill = _NO_FILL
                    cell.value = None
            else:
                is_weekend = date(cfg.anno, month_num, day).weekday() >= 5
                fill = _WEEKEND_FILL if is_weekend else _NO_FILL
                for r in range(_ROW_ORE_PROGETTO, _ROW_ORE_ALTRO + 1):
                    ws[f"{col_letter}{r}"].fill = fill

                # Ore: solo per mesi inclusi nella configurazione
                if month_num in cfg.mesi and day in day_to_ore:
                    ore_prog, ore_ord = day_to_ore[day]
                    ws[f"{col_letter}{_ROW_ORE_PROGETTO}"].value = ore_prog or None
                    ws[f"{col_letter}{_ROW_ORE_ORDINARIE}"].value = ore_ord or None


def _aggiorna_riferimenti(formula: str, rename_map: dict[str, str]) -> str:
    """Sostituisce i nomi foglio nelle stringhe formula del Riassuntivo."""
    for old_name, new_name in rename_map.items():
        # Nomi con spazi sono tra singoli apici nelle formule Excel
        formula = formula.replace(f"'{old_name}'", f"'{new_name}'")
        # Caso senza apici (sheet senza spazi, non presente ma per robustezza)
        formula = formula.replace(f"{old_name}!", f"{new_name}!")
    return formula
